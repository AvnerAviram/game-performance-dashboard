#!/usr/bin/env python3
"""
Continue Theme Validation from Checkpoint (307/520)
Processes remaining 213 games from Data Download Theme (4).xlsx
"""

import openpyxl
import json
import os
from datetime import datetime

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = '/Users/avner/Downloads/Data Download Theme (4).xlsx'
GAMES_MASTER_PATH = os.path.join(SCRIPT_DIR, '..', 'game_analytics_export', 'data', 'games_master.json')
CHECKPOINT_PATH = os.path.join(SCRIPT_DIR, 'theme_validation_checkpoint.json')
PROGRESS_LOG_PATH = os.path.join(SCRIPT_DIR, 'theme_validation_progress.log')

def load_excel_data():
    """Load and parse the Excel file"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    print(f"Excel file loaded: {ws.max_row} rows, {ws.max_column} columns")
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    
    # Parse all rows (skip header row 1 and grand total last row)
    games_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1, values_only=True):
        # Check if this is a valid data row (has an index or game name)
        if row[2] or row[3]:  # Index or Game Name column
            game_dict = dict(zip(headers, row))
            # Clean game name
            game_name = game_dict.get('Game Name', '')
            if game_name and isinstance(game_name, str) and game_name.strip():
                games_data.append(game_dict)
    
    return games_data

def load_games_master():
    """Load the games master JSON"""
    with open(GAMES_MASTER_PATH, 'r') as f:
        data = json.load(f)
    return data

def load_checkpoint():
    """Load validation checkpoint if exists"""
    if os.path.exists(CHECKPOINT_PATH):
        with open(CHECKPOINT_PATH, 'r') as f:
            return json.load(f)
    return {'validated_count': 0, 'validated_games': [], 'last_updated': None}

def save_checkpoint(checkpoint_data):
    """Save validation checkpoint"""
    checkpoint_data['last_updated'] = datetime.utcnow().isoformat() + 'Z'
    with open(CHECKPOINT_PATH, 'w') as f:
        json.dump(checkpoint_data, f, indent=2)

def log_progress(message):
    """Log progress to file and console"""
    timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    log_message = f"[{timestamp}] {message}"
    print(log_message)
    with open(PROGRESS_LOG_PATH, 'a') as f:
        f.write(log_message + '\n')

def find_game_in_master(game_name, games_master):
    """Find a game in the master list (case-insensitive)"""
    game_name_lower = game_name.lower().strip()
    for game in games_master['games']:
        if game['name'].lower().strip() == game_name_lower:
            return game
    return None

def update_game_theme(game, theme_data):
    """Update game theme information from Excel row"""
    if 'theme' not in game:
        game['theme'] = {}
    
    # Update theme fields based on Excel columns
    if 'Theme Primary' in theme_data and theme_data['Theme Primary']:
        game['theme']['primary'] = theme_data['Theme Primary']
    
    if 'Theme Consolidated' in theme_data and theme_data['Theme Consolidated']:
        game['theme']['consolidated'] = theme_data['Theme Consolidated']
    
    # Mark as validated
    if 'classification' not in game:
        game['classification'] = {}
    game['classification']['theme_validated'] = True
    game['classification']['theme_validation_date'] = datetime.utcnow().isoformat() + 'Z'
    
    return game

def main():
    log_progress("=" * 80)
    log_progress("Starting Theme Validation Continuation (Checkpoint: 307/520)")
    log_progress("=" * 80)
    
    # Load data
    log_progress("Loading Excel data...")
    excel_games = load_excel_data()
    total_games = len(excel_games)
    log_progress(f"Loaded {total_games} games from Excel")
    
    log_progress("Loading games master...")
    games_master = load_games_master()
    log_progress(f"Loaded {len(games_master['games'])} games from master")
    
    # Load checkpoint
    checkpoint = load_checkpoint()
    start_index = checkpoint.get('validated_count', 0)
    
    if start_index == 0:
        # If no checkpoint, assume 307 games were already validated
        start_index = 307
        log_progress(f"No checkpoint found. Starting from row {start_index + 1}")
    else:
        log_progress(f"Resuming from checkpoint: {start_index}/{total_games} games validated")
    
    # Process remaining games
    updated_count = 0
    not_found_count = 0
    skipped_count = 0
    
    for i in range(start_index, total_games):
        excel_row = excel_games[i]
        game_name = excel_row.get('Game Name') or excel_row.get('Name') or excel_row.get('name')
        
        if not game_name:
            log_progress(f"Row {i+1}: Skipping - no game name")
            skipped_count += 1
            continue
        
        # Find game in master
        master_game = find_game_in_master(game_name, games_master)
        
        if not master_game:
            log_progress(f"Row {i+1}/{total_games}: '{game_name}' NOT FOUND in master")
            not_found_count += 1
            continue
        
        # Update theme
        update_game_theme(master_game, excel_row)
        updated_count += 1
        
        if (i + 1) % 10 == 0:
            log_progress(f"Progress: {i+1}/{total_games} processed ({updated_count} updated, {not_found_count} not found)")
            
            # Save checkpoint every 10 games
            checkpoint['validated_count'] = i + 1
            checkpoint['validated_games'].append(game_name)
            save_checkpoint(checkpoint)
    
    # Save updated master file
    log_progress("Saving updated games master...")
    with open(GAMES_MASTER_PATH, 'w') as f:
        json.dump(games_master, f, indent=2)
    
    # Final report
    log_progress("=" * 80)
    log_progress("VALIDATION COMPLETE!")
    log_progress("=" * 80)
    log_progress(f"Total games in Excel: {total_games}")
    log_progress(f"Started from index: {start_index}")
    log_progress(f"Games processed: {total_games - start_index}")
    log_progress(f"Games updated: {updated_count}")
    log_progress(f"Games not found: {not_found_count}")
    log_progress(f"Games skipped: {skipped_count}")
    log_progress(f"Final count: {start_index + updated_count}/{total_games}")
    log_progress("=" * 80)
    
    # Save final checkpoint
    checkpoint['validated_count'] = total_games
    checkpoint['status'] = 'completed'
    checkpoint['summary'] = {
        'total': total_games,
        'updated': updated_count,
        'not_found': not_found_count,
        'skipped': skipped_count
    }
    save_checkpoint(checkpoint)
    
    print(f"\nCheckpoint saved to: {CHECKPOINT_PATH}")
    print(f"Progress log saved to: {PROGRESS_LOG_PATH}")
    print(f"Updated games master: {GAMES_MASTER_PATH}")

if __name__ == '__main__':
    main()
